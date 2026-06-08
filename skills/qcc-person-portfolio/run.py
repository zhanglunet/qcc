#!/usr/bin/env python3
"""
qcc-person-portfolio — 给一个自然人,反查他作为董监高/法代/控股/关联的所有公司,导 xlsx。

Usage:
    # 模式 A:已知 anchor USCC(最快)
    python run.py --person "雷军" --anchor 91110108551385082Q

    # 模式 B:让脚本自动 anchor(传公司名)
    python run.py --person "雷军" --anchor-by-name "小米科技"

    # 指定输出
    python run.py --person "雷军" --anchor 91110108551385082Q --out ./leijun.xlsx

依赖:
    cd qcc/python && source .venv/bin/activate
    pip install -e ".[skills]"   # 装 openpyxl

环境:
    QCC_API_KEY 必须可读 — 脚本会按以下顺序找 .env:
      1. 当前工作目录 ./.env
      2. 仓库根 ../../.env
      3. 系统环境变量
"""
from __future__ import annotations

import argparse
import asyncio
import os
import re
import sys
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv

# Resolve .env from CWD or from the qcc repo root (2 levels above this script)
HERE = Path(__file__).resolve().parent
for candidate in (Path.cwd() / ".env", HERE.parent.parent / ".env"):
    if candidate.exists():
        load_dotenv(candidate)
        break
else:
    load_dotenv()  # fall back to default search

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl missing — install with: pip install -e '.[skills]'  (from qcc/python)")

from qcc_client import QccClient, Server   # noqa: E402

USCC_RE = re.compile(r"^[0-9A-HJ-NPQRTUWXY]{18}$")

TOOLS = {
    "current_positions":     ("get_executive_positions",                  "当前董监高任职", "董监高-在外任职信息"),
    "historical_positions":  ("get_executive_historical_positions",       "历史董监高任职", "董监高-历史在外任职信息"),
    "current_legal_rep":     ("get_executive_legal_rep_roles",            "当前法定代表人", "董监高-担任法定代表人信息"),
    "historical_legal_rep":  ("get_executive_historical_legal_rep_roles", "历史法定代表人", "董监高-历史担任法定代表人信息"),
    "controlled":            ("get_executive_controlled_companies",       "控制企业",       "董监高-控制企业信息"),
    "related":               ("get_executive_related_companies",          "关联企业",       "董监高-全部关联企业信息"),
}

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ZEBRA_FILL  = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


# ---------- anchor resolution ----------

async def resolve_anchor(client: QccClient, anchor_arg: str, person: str, anchor_name: str | None) -> str:
    if anchor_arg and USCC_RE.match(anchor_arg):
        return anchor_arg

    if not anchor_name:
        sys.exit("anchor missing: provide --anchor <USCC> or --anchor-by-name <company-name>")

    print(f"Resolving anchor via name lookup: {anchor_name!r} ...", file=sys.stderr)
    result = await client.call(Server.COMPANY, "get_company_by_query", {"searchKey": anchor_name})

    status = result.get("匹配结果") if isinstance(result, dict) else None
    if status == "唯一精确匹配":
        info = result.get("企业信息", {})
        uscc = info.get("统一社会信用代码")
        legal_reps = info.get("法定代表人") or info.get("法定代表人名称") or []
        if person in str(legal_reps):
            print(f"  ✓ anchored: {info.get('企业名称')} ({uscc}), 法代 = {legal_reps}", file=sys.stderr)
        else:
            print(f"  ⚠️ anchored: {info.get('企业名称')} ({uscc}), but 法代={legal_reps} does NOT contain {person!r}. Continuing anyway.", file=sys.stderr)
        return uscc

    if status == "多候选":
        candidates = result.get("企业信息", [])
        preferred = [c for c in candidates if person in str(c.get("法定代表人名称", []))]
        if len(preferred) == 1:
            c = preferred[0]
            print(f"  ✓ anchored (unique person-legal-rep match): {c['企业名称']} ({c['统一社会信用代码']})", file=sys.stderr)
            return c["统一社会信用代码"]
        if not preferred:
            sys.exit(f"no candidates where 法代 = {person!r}; try a different --anchor-by-name")
        print(f"\nMultiple candidates have {person!r} as 法代; pick the right anchor manually:", file=sys.stderr)
        for i, c in enumerate(preferred, 1):
            print(f"  [{i}] {c['企业名称']} ({c['统一社会信用代码']}) - 状态:{c.get('状态')}", file=sys.stderr)
        sys.exit("re-run with --anchor <USCC> of the chosen one")

    sys.exit(f"no match for {anchor_name!r}; check spelling")


# ---------- fetch + normalize ----------

async def fetch_all(client: QccClient, anchor_uscc: str, person: str) -> dict:
    args = {"searchKey": anchor_uscc, "personName": person}
    tasks = {k: client.call(Server.EXECUTIVE, tool, args) for k, (tool, _, _) in TOOLS.items()}
    results = await asyncio.gather(*tasks.values(), return_exceptions=True)
    return dict(zip(tasks.keys(), results))


def normalize(key: str, payload) -> list[dict]:
    if isinstance(payload, Exception):
        return [{"_error": str(payload)}]
    if not isinstance(payload, dict):
        return []
    _, _, field = TOOLS[key]
    records = payload.get(field, []) or []
    out = []
    for r in records:
        role = r.get("职位") or r.get("角色") or ""
        if isinstance(role, list):
            role = " / ".join(str(x) for x in role)
        out.append({
            "企业名称":  r.get("企业名称") or "",
            "状态":      r.get("状态") or "",
            "职位 / 角色": role,
            "持股比例":  r.get("持股比例") or r.get("投资比例") or "",
            "任职/担任起止时间": r.get("任职起止时间") or r.get("担任法定代表人起止时间") or "",
            "注册资本":  r.get("注册资本") or "",
            "地区":      r.get("地区") or r.get("所属地区") or "",
            "行业":      r.get("行业") or r.get("所属行业") or "",
            "成立日期":  r.get("成立日期") or "",
        })
    return out


def build_master(by_source: dict[str, list[dict]]) -> list[dict]:
    bucket: dict[str, dict] = {}
    for key, rows in by_source.items():
        _, label, _ = TOOLS[key]
        for r in rows:
            name = r.get("企业名称", "")
            if not name or name.startswith("_"):
                continue
            slot = bucket.setdefault(name, {
                "企业名称": name,
                "出现的来源": set(),
                "职位 / 角色 汇总": set(),
                "持股比例(最新一次出现)": "",
                "最新一次状态": "",
                "最新一次地区": "",
                "最新一次行业": "",
                "成立日期": "",
                "注册资本": "",
            })
            slot["出现的来源"].add(label)
            v = r.get("职位 / 角色", "")
            if v:
                slot["职位 / 角色 汇总"].add(str(v))
            for k_target, k_source in [
                ("持股比例(最新一次出现)", "持股比例"),
                ("最新一次状态", "状态"),
                ("最新一次地区", "地区"),
                ("最新一次行业", "行业"),
                ("成立日期", "成立日期"),
                ("注册资本", "注册资本"),
            ]:
                if not slot[k_target] and r.get(k_source):
                    slot[k_target] = r[k_source]
    master = []
    for name in sorted(bucket):
        s = bucket[name]
        s["出现的来源"] = " / ".join(sorted(s["出现的来源"]))
        s["职位 / 角色 汇总"] = " / ".join(sorted(s["职位 / 角色 汇总"]))
        master.append(s)
    return master


# ---------- xlsx writer ----------

def style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def autosize(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_w = 0
        for cell in ws[letter]:
            v = str(cell.value or "")
            w = sum(2 if ord(c) > 127 else 1 for c in v)
            max_w = max(max_w, w)
        ws.column_dimensions[letter].width = min(max_w + 2, 60)


def write_sheet(ws, rows: list[dict]) -> None:
    if not rows:
        ws.append(["(无记录)"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for i, r in enumerate(rows, start=2):
        ws.append([r.get(h, "") for h in headers])
        if i % 2 == 0:
            for c in range(1, len(headers) + 1):
                ws.cell(row=i, column=c).fill = ZEBRA_FILL
    style_header(ws, len(headers))
    autosize(ws)


def write_xlsx(out_path: Path, person: str, anchor_uscc: str, by_source: dict[str, list[dict]], master: list[dict], summary_capped: dict[str, str]) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    info = wb.create_sheet("说明", 0)
    info.append(["字段", "值"])
    info.append(["查询人物", person])
    info.append(["锚定 USCC", anchor_uscc])
    info.append(["数据源", "agent.qcc.com (企查查 Agent MCP, executive server)"])
    info.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    info.append(["数据时效", "T+0"])
    info.append([])
    info.append(["Sheet 名", "记录数", "工具名", "备注"])
    info.append(["汇总(去重)", len(master), "(合并)",
                 "按企业名称去重,标记每家公司出现在哪些来源中"])
    for key in TOOLS:
        tool, label, _ = TOOLS[key]
        capped = summary_capped.get(key, "")
        info.append([label, len(by_source[key]), tool, capped])
    info.append([])
    info.append(["⚠️ 提示", "「董监高」= 董事 / 监事 / 高级管理人员;法代另单列便于核对"])
    info.append(["⚠️ 提示", "境外/港股/美股开曼 VIE 主体不在 QCC 工商数据内"])
    style_header(info, 4)
    autosize(info)

    master_ws = wb.create_sheet("汇总(去重)")
    write_sheet(master_ws, master)

    for key, (_, label, _) in TOOLS.items():
        ws = wb.create_sheet(label)
        write_sheet(ws, by_source[key])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ---------- CLI ----------

def parse_args():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--person", required=True, help="中文姓名")
    ap.add_argument("--anchor", default="", help="已知 anchor USCC(18 位)")
    ap.add_argument("--anchor-by-name", default="", dest="anchor_by_name",
                    help="用公司名让脚本自动 anchor")
    ap.add_argument("--out", default="", help="xlsx 输出路径(默认 ./{person}_companies.xlsx)")
    return ap.parse_args()


async def amain():
    args = parse_args()
    out_path = Path(args.out) if args.out else Path(f"./{args.person}_companies.xlsx")

    client = QccClient.from_env()

    anchor_uscc = await resolve_anchor(client, args.anchor, args.person, args.anchor_by_name)

    print(f"\nFetching for {args.person} (anchor={anchor_uscc}) ...")
    raw = await fetch_all(client, anchor_uscc, args.person)

    by_source: dict[str, list[dict]] = {}
    summary_capped: dict[str, str] = {}
    for k, payload in raw.items():
        rows = normalize(k, payload)
        by_source[k] = rows
        if isinstance(payload, dict):
            hint = payload.get("提示", "") or ""
            summary_line = payload.get("摘要", "") or ""
            m = re.search(r"共有\s*(\d+)\s*条", summary_line)
            total_in_summary = int(m.group(1)) if m else None
            if total_in_summary and total_in_summary > len(rows):
                summary_capped[k] = f"⚠️ 截断:原 {total_in_summary} 条,QCC 仅返回前 {len(rows)} 条"
            elif "前" in hint and "条" in hint:
                summary_capped[k] = f"⚠️ {hint}"

    for k, rows in by_source.items():
        _, label, _ = TOOLS[k]
        print(f"  {label}: {len(rows)} {summary_capped.get(k, '')}")

    master = build_master(by_source)
    print(f"\nUnique companies after dedupe: {len(master)}")

    write_xlsx(out_path, args.person, anchor_uscc, by_source, master, summary_capped)
    print(f"\n✓ Written: {out_path}")
    print(f"  size: {out_path.stat().st_size:,} bytes")


if __name__ == "__main__":
    asyncio.run(amain())
